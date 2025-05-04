import re
from openpyxl import load_workbook
from pptx import Presentation
from typing  import List, Dict
from pprint import pprint, pformat
import sys

from domain.iopen_document import IOpenDocument
from domain.worker_class import Worker
from domain.queue import MetadataDoc, MetadataXls, MetadataPpt
from infrastructure.generic_logger import GenericLogger

class IOpenAndUpdateDocument(IOpenDocument):
    def __init__(self, document_path: str, 
                 worker: Worker, 
                 paragraph_start_min_word_numbers: int, paragraph_start_min_word_length: int,  
                 logger: GenericLogger):
        self.document_path = document_path
        self.worker = worker
        self.logger = logger
        self.paragraph_start_min_word_length = paragraph_start_min_word_length
        self.paragraph_start_min_word_numbers = paragraph_start_min_word_numbers
        self.worker: Worker = worker

    def is_paragraph(self, text: str):
        minwords: int = int(self.paragraph_start_min_word_numbers)  # - 1 if int(self.paragraph_start_min_word_numbers) - 1 >= 0 else 0
        regexp: str = f'(\\w{{{self.paragraph_start_min_word_length},}}\\b\\s+){{{minwords},}}\w{{{self.paragraph_start_min_word_length},}}\\b'
        paragraph_found: bool = (re.search(regexp, text) is not None)
        self.logger.log_trace(f"Checking for paragraph for {text} with regexp: {regexp} is {paragraph_found}")
        return paragraph_found
        
    def save(self, filename: str) -> None:
        self.document.save(filename)
        self.logger.log_info(f"Saved final document as {filename}")


class OpenXLSDocument(IOpenAndUpdateDocument):
    def __init__(self, document_path: str, 
                 worker: Worker, 
                 paragraph_start_min_word_numbers: int,
                 paragraph_start_min_word_length: int, 
                 logger: GenericLogger):
        super().__init__(document_path, 
                         worker, 
                         paragraph_start_min_word_numbers, paragraph_start_min_word_length, 
                         logger)
        self.document =  load_workbook(document_path)
    
    def __fill_tasks(self, xls_obj: any):
        for ws in xls_obj.worksheets:

            for row in range(1,ws.max_row + 1):
                for col in range(1,ws.max_column + 1):
                    current_text = str(ws.cell(row,col).value)
                    if current_text is not None and current_text != 'None' and self.is_paragraph(current_text): 
                        self.worker.add_work_element(MetadataXls(ws.cell(row,col)))

    def process(self):
        self.__fill_tasks(self.document)
        self.worker.process_all()

